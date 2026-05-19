import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
import numpy as np

# ------------------------------
# 1. 定义各工作表的数据结构
# ------------------------------

# Q1_结果：关键指标筛选 + 体质贡献度
q1_indicators = pd.DataFrame({
    "指标名称": ["总胆固醇(TC)", "甘油三酯(TG)", "低密度脂蛋白(LDL-C)", "高密度脂蛋白(HDL-C)",
                "血糖", "血尿酸", "BMI", "ADL总分", "IADL总分", "活动量表总分"],
    "与痰湿积分的相关系数": [None] * 10,
    "与高血脂标签的相关系数": [None] * 10,
    "是否入选关键指标": ["待填"] * 10,
    "筛选逻辑说明": ["待填"] * 10
})

# 九种体质贡献度
q1_constitutions = pd.DataFrame({
    "体质类型": ["平和质", "气虚质", "阳虚质", "阴虚质", "痰湿质", "湿热质", "血瘀质", "气郁质", "特禀质"],
    "体质标签": [1,2,3,4,5,6,7,8,9],
    "确诊人群平均积分": [None]*9,
    "未确诊人群平均积分": [None]*9,
    "风险贡献度(OR/系数)": [None]*9,
    "贡献度排序": [None]*9
})

# Q2_风险分级：每个样本的风险等级及特征组合
# 这里只生成示例行（如前5个样本），实际应用时可读取全部1000条数据
q2_risk = pd.DataFrame({
    "样本ID": [1,2,3,4,5],
    "痰湿积分": [None]*5,
    "活动量表总分": [None]*5,
    "血脂异常标签(0/1)": [None]*5,
    "风险等级": ["待填"]*5,
    "阈值依据说明": ["待填"]*5,
    "是否为高风险痰湿体质": ["待填"]*5,
    "核心特征组合": ["待填"]*5
})

# Q2_特征组合统计（汇总高风险人群的常见模式）
q2_patterns = pd.DataFrame({
    "特征组合模式": ["痰湿积分≥80 + 活动评分<40 + TG≥2.0",
                    "痰湿积分≥70 + 活动评分<50 + LDL-C≥3.5",
                    "痰湿积分≥60 + 活动评分<60 + 确诊高血脂"],
    "出现频次": [None]*3,
    "占比(%)": [None]*3
})

# Q3_干预方案：针对痰湿体质(体质=5)的所有样本
# 示例行（ID=1,2,3 及两个其他示例）
q3_intervention = pd.DataFrame({
    "样本ID": [1,2,3,100,200],
    "年龄组": [None]*5,
    "活动量表总分": [None]*5,
    "初始痰湿积分": [None]*5,
    "调理分级(1/2/3)": [None]*5,
    "活动干预强度(1/2/3)": [None]*5,
    "训练频率(次/周)": [None]*5,
    "总成本(6个月)": [None]*5,
    "预期痰湿积分下降": [None]*5,
    "是否为最优方案": ["待填"]*5,
    "优化逻辑说明": ["待填"]*5
})

# Q3_样本1_2_3方案：单独列出ID=1,2,3的最优方案
q3_sample123 = pd.DataFrame({
    "样本ID": [1,2,3],
    "年龄组": [None]*3,
    "活动评分": [None]*3,
    "初始痰湿积分": [None]*3,
    "最优调理分级": [None]*3,
    "最优活动强度": [None]*3,
    "最优频率(次/周)": [None]*3,
    "总成本(元)": [None]*3,
    "预期积分下降": [None]*3
})

# 参数与假设说明（作为文本表格）
params_data = {
    "类别": ["调理分级成本", "调理分级成本", "调理分级成本",
             "活动强度单位成本", "活动强度单位成本", "活动强度单位成本",
             "年龄约束", "年龄约束", "年龄约束",
             "活动评分约束", "活动评分约束", "活动评分约束",
             "积分下降规律", "积分下降规律", "成本上限"],
    "参数/规则": ["基础调理(1级)", "中度调理(2级)", "强化调理(3级)",
                 "1级强度(10min/次)", "2级强度(20min/次)", "3级强度(30min/次)",
                 "40-59岁", "60-79岁", "80-89岁",
                 "活动总分<40", "40≤总分<60", "总分≥60",
                 "每周训练5次，每提升一级强度", "固定强度下每周增加1次", "单人6个月总成本"],
    "取值/说明": ["30元/月", "80元/月", "130元/月",
                 "3元/次", "5元/次", "8元/次",
                 "可选1/2/3级", "可选1/2级", "仅可选1级",
                 "仅1级强度", "可选1/2级", "可选1/2/3级",
                 "每月痰湿积分下降3%", "每月痰湿积分下降1%", "≤2000元"]
}
q4_params = pd.DataFrame(params_data)

# ------------------------------
# 2. 写入Excel文件（多sheet）
# ------------------------------
output_file = "MathorCup_C_Result_Template.xlsx"

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    q1_indicators.to_excel(writer, sheet_name="Q1_关键指标筛选", index=False)
    q1_constitutions.to_excel(writer, sheet_name="Q1_体质贡献度", index=False)
    q2_risk.to_excel(writer, sheet_name="Q2_风险分级", index=False)
    q2_patterns.to_excel(writer, sheet_name="Q2_特征组合统计", index=False)
    q3_intervention.to_excel(writer, sheet_name="Q3_干预方案_全体痰湿体质", index=False)
    q3_sample123.to_excel(writer, sheet_name="Q3_样本1_2_3最优方案", index=False)
    q4_params.to_excel(writer, sheet_name="参数与假设说明", index=False)

# ------------------------------
# 3. 美化格式（列宽、对齐、标题颜色等）
# ------------------------------
wb = load_workbook(output_file)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # 列宽自动调整（简单设置）
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)  # 最大宽度30
        ws.column_dimensions[col_letter].width = adjusted_width
    # 标题行样式
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

# 额外：在“参数与假设说明”中增加注释说明
ws_params = wb["参数与假设说明"]
ws_params.append(["", "", "注：以上为模型核心参数，实际应用时可根据数据调整"])
ws_params.cell(row=len(params_data)+2, column=1).font = Font(italic=True)

wb.save(output_file)
print(f"Excel模板已生成：{output_file}")
print("请根据实际计算结果填充空白单元格，特别是相关系数、风险等级、最优方案等。")